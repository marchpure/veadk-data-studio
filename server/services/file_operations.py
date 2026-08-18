"""File dataset operations service for CSV, Excel, Parquet, JSON files."""

from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import logging
import os
import re
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb  # type: ignore
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from server.config.storage import dataset_directory
from server.models.files import File
from server.repositories.files import FileRepository
from server.services.dataset import DatasetService
from server.services.dataset_storage import DatasetStorageService, StoredFileMetadata
from server.services.duckdb_service import DuckDBFileDescriptor, DuckDBService

try:
    import chardet

    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

logger = logging.getLogger(__name__)


_ARTIFACT_LOCKS: dict[str, asyncio.Lock] = {}
_ARTIFACT_LOCKS_MUTEX = asyncio.Lock()


@dataclass(slots=True)
class ColumnarArtifact:
    """Metadata describing a materialized columnar copy of a dataset file."""

    storage: StoredFileMetadata
    format: str
    row_count: int | None


class DataFrameFileService:
    """Unified service for DuckDB-backed file datasets (CSV, Excel, Parquet, JSON)."""

    @staticmethod
    def get_cached_schema(dataset) -> dict[str, Any] | None:
        """
        Get cached schema from dataset if available.

        Args:
            dataset: Dataset model instance

        Returns:
            Cached schema dict or None if not available/invalid
        """
        try:
            if dataset.schema_cache:
                try:
                    return json.loads(dataset.schema_cache)
                except json.JSONDecodeError:
                    logger.error(
                        f"Failed to parse cached schema for dataset {dataset.id}",
                        exc_info=False,
                    )
                    return None
            return None
        except Exception as e:
            logger.error(
                f"Failed to get cached schema: {str(e)}",
                exc_info=True,
            )
            return None

    @staticmethod
    async def save_schema_cache(
        session: AsyncSession,
        dataset,
        schema_data: dict[str, Any],
    ) -> None:
        """
        Save schema to dataset cache.

        Args:
            session: Database session
            dataset: Dataset model instance
            schema_data: Schema dictionary to cache
        """
        try:
            dataset.schema_cache = json.dumps(schema_data)
            dataset.schema_updated_at = datetime.utcnow()
            session.add(dataset)
            await session.commit()
            await session.refresh(dataset)
            logger.info(f"Successfully cached schema for dataset {dataset.id}")
        except Exception as e:
            logger.error(
                f"Failed to save schema cache for dataset {dataset.id}: {str(e)}",
                exc_info=True,
            )
            raise

    @staticmethod
    def _alias_from_filename(filename: str) -> str:
        """Derive a stable SQL-safe alias from a filename (without extension)."""
        import re

        # Remove extension
        base = filename.rsplit(".", 1)[0] if "." in filename else filename
        # Replace special characters with underscores for SQL compatibility
        # Keep only alphanumeric and underscore, replace everything else
        sanitized = re.sub(r"[^a-zA-Z0-9_]", "_", base)
        # Ensure it doesn't start with a number (SQL identifiers can't start with digits)
        if sanitized and sanitized[0].isdigit():
            sanitized = f"t_{sanitized}"
        # Ensure it's not empty
        return sanitized if sanitized else "table"

    @staticmethod
    def _next_alias(alias_counts: dict[str, int], candidate: str) -> str:
        """Generate a deduplicated alias name."""
        counter = alias_counts.get(candidate, 0)
        alias_counts[candidate] = counter + 1
        return candidate if counter == 0 else f"{candidate}_{counter}"

    @staticmethod
    def _dataset_root(dataset) -> Path:
        """Return the root directory for the dataset, updating storage_path if missing."""
        if dataset.storage_path:
            current = Path(dataset.storage_path).expanduser()
            if current.exists():
                return current.resolve()
            logger.warning(
                "Dataset %s storage path %s is missing; remapping to current dataset directory",
                getattr(dataset, "id", "<unknown>"),
                dataset.storage_path,
            )

        root = dataset_directory(str(dataset.id))
        dataset.storage_path = str(root)

        duckdb_dir = root / "duckdb"
        duckdb_dir.mkdir(parents=True, exist_ok=True)

        existing_duckdb = getattr(dataset, "duckdb_path", None)
        if existing_duckdb:
            duckdb_path = Path(existing_duckdb).expanduser()
            if duckdb_path.exists():
                dataset.duckdb_path = str(duckdb_path.resolve())
            else:
                filename = duckdb_path.name if duckdb_path.name not in {"", "."} else "dataset.duckdb"
                dataset.duckdb_path = str((duckdb_dir / filename).resolve())
        else:
            dataset.duckdb_path = str((duckdb_dir / "dataset.duckdb").resolve())

        return root.resolve()

    @staticmethod
    async def _get_artifact_lock(file_id: str) -> asyncio.Lock:
        """Retrieve (or create) an asyncio lock scoped to a file id."""
        async with _ARTIFACT_LOCKS_MUTEX:
            lock = _ARTIFACT_LOCKS.get(file_id)
            if lock is None:
                lock = asyncio.Lock()
                _ARTIFACT_LOCKS[file_id] = lock
        return lock

    @staticmethod
    async def _ensure_file_materialized(
        session: AsyncSession,
        dataset,
        file: File,
    ) -> Path:
        """
        Ensure a SQLAlchemy File record has an accessible on-disk path.

        For legacy rows that only store blob content this writes the content to
        the dataset's raw directory and updates the record with the new path.
        """
        base_dir = DataFrameFileService._dataset_root(dataset)

        if file.storage_path:
            candidate = (base_dir / file.storage_path).resolve()
            if candidate.exists():
                return candidate

        if file.content is not None:
            metadata = await DatasetStorageService.save_bytes(
                dataset_id=str(dataset.id),
                filename=file.name,
                data=file.content,
            )
            file.storage_path = str(metadata.relative_path)
            file.size = metadata.size
            file.checksum = metadata.checksum
            session.add(file)
            await session.commit()
            await session.refresh(file)
            return metadata.absolute_path

        raise FileNotFoundError(f"File {file.id} has no storage path or inline content")

    @staticmethod
    async def _build_duckdb_descriptors(
        session: AsyncSession,
        dataset,
    ) -> list[DuckDBFileDescriptor]:
        """Materialize dataset files and build DuckDB descriptors."""
        supported_types = set(DuckDBService._SQL_READERS.keys())
        descriptors: list[DuckDBFileDescriptor] = []
        alias_counts: dict[str, int] = {}

        dataset_root = DataFrameFileService._dataset_root(dataset)
        excel_supported = DuckDBService.excel_extension_available()

        for file in dataset.files:
            original_type = (file.type or "").lower()
            effective_type = (file.optimized_format or original_type).lower()

            if effective_type not in supported_types:
                supported_list = ", ".join(sorted(supported_types))
                raise ValueError(
                    f"DuckDB query support not implemented for file type '{effective_type or file.type}'. "
                    f"Supported types: {supported_list}."
                )

            file_path: Path | None = None

            if file.optimized_storage_path:
                optimized_candidate = (dataset_root / file.optimized_storage_path).resolve()
                if optimized_candidate.exists():
                    file_path = optimized_candidate
                else:
                    logger.warning(
                        "Optimized artifact missing for file %s (dataset %s). Falling back to raw file.",
                        file.id,
                        dataset.id,
                    )

            if file_path is None:
                file_path = await DataFrameFileService._ensure_file_materialized(session, dataset, file)
                effective_type = original_type

            base_alias = DataFrameFileService._alias_from_filename(file.name).strip()
            if not base_alias:
                base_alias = f"file_{file.id or len(descriptors) + 1}"

            if effective_type == "excel":
                sheet_names = DataFrameFileService.get_excel_sheets(file_path)
                if not sheet_names:
                    raise ValueError(f"Excel file '{file.name}' contains no sheets.")

                multi_sheet = len(sheet_names) > 1
                for idx, sheet_name in enumerate(sheet_names):
                    sheet_alias = DataFrameFileService._alias_from_filename(sheet_name).strip() or f"sheet_{idx + 1}"
                    candidate_alias = base_alias if not multi_sheet else f"{base_alias}__{sheet_alias}"
                    if not candidate_alias:
                        candidate_alias = f"{base_alias}_sheet_{idx + 1}"

                    alias = DataFrameFileService._next_alias(alias_counts, candidate_alias)

                    # Check file extension - only .xlsx supports native DuckDB read_xlsx()
                    # .xls (Excel 97-2003) must use CSV conversion fallback
                    file_extension = file_path.suffix.lower()
                    use_native_excel = excel_supported and file_extension == ".xlsx"

                    if file_extension == ".xls":
                        logger.info(
                            "File %s is .xls format (Excel 97-2003). Using CSV conversion fallback. "
                            "Native DuckDB Excel extension only supports .xlsx format.",
                            file.name,
                        )

                    if use_native_excel:
                        descriptors.append(
                            DuckDBFileDescriptor(
                                alias=alias,
                                path=file_path,
                                file_type="excel",
                                filename=file.name,
                                sheet_name=sheet_name,
                                reader_options={"sheet": sheet_name},
                            )
                        )
                    else:
                        sheet_path = await DataFrameFileService._ensure_excel_sheet_parquet(
                            dataset=dataset,
                            file_record=file,
                            source_path=file_path,
                            sheet_name=sheet_name,
                        )
                        descriptors.append(
                            DuckDBFileDescriptor(
                                alias=alias,
                                path=sheet_path,
                                file_type="parquet",
                                filename=file.name,
                                sheet_name=sheet_name,
                                reader_options={"sheet": sheet_name, "source_type": "excel"},
                            )
                        )
            else:
                alias = DataFrameFileService._next_alias(alias_counts, base_alias)

                descriptors.append(
                    DuckDBFileDescriptor(
                        alias=alias,
                        path=file_path,
                        file_type=effective_type,
                        filename=file.name,
                    )
                )

        return descriptors

    @staticmethod
    async def _collect_duckdb_single_file_schema(
        descriptor: DuckDBFileDescriptor,
        sample_rows: int,
    ) -> dict[str, Any]:
        """Collect schema and samples for a single file using DuckDB."""
        duckdb_schema = await DuckDBService.collect_schema([descriptor], sample_rows=sample_rows)
        alias = descriptor.alias

        table_info = dict(duckdb_schema.get("schema", {}).get(alias, {}))
        sample_data = list(duckdb_schema.get("sample_data", {}).get(alias, []))

        if table_info and "filename" not in table_info:
            table_info["filename"] = descriptor.filename
        if table_info and "file_type" not in table_info:
            table_info["file_type"] = descriptor.file_type

        return {
            "datasource_type": descriptor.file_type,
            "datasource_name": descriptor.filename,
            "filename": descriptor.filename,
            "columns": table_info.get("columns", []),
            "row_count": table_info.get("row_count", 0),
            "sample_data": sample_data,
            "schema": {
                alias: table_info,
            },
        }

    @staticmethod
    def _materialize_temp_file(file_content: bytes | str, file_type: str) -> Path:
        """Write file content to a temporary file so DuckDB can read it."""
        suffix = f".{file_type}" if file_type else ""
        data = file_content if isinstance(file_content, bytes) else str(file_content).encode("utf-8")

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        try:
            temp_file.write(data)
            temp_file.flush()
            return Path(temp_file.name)
        finally:
            temp_file.close()

    @staticmethod
    def _open_excel_workbook(file_source: bytes | str | Path):
        """Open an Excel workbook from either bytes or a filesystem path."""
        if isinstance(file_source, (str, Path)):
            return load_workbook(filename=str(file_source), read_only=True, data_only=True)
        if isinstance(file_source, bytes):
            return load_workbook(filename=io.BytesIO(file_source), read_only=True, data_only=True)
        if isinstance(file_source, io.BytesIO):
            return load_workbook(filename=file_source, read_only=True, data_only=True)
        raise ValueError("Unsupported Excel source type")

    @staticmethod
    def _compute_file_checksum(path: Path, chunk_size: int = 8 * 1024 * 1024) -> str:
        """Compute SHA256 checksum of a file without loading it fully into memory."""
        digest = hashlib.sha256()
        with path.open("rb") as infile:
            while True:
                chunk = infile.read(chunk_size)
                if not chunk:
                    break
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _materialize_descriptor_to_parquet_sync(
        dataset_id: str,
        descriptor: DuckDBFileDescriptor,
        existing_relative_path: str | None = None,
    ) -> ColumnarArtifact:
        """Materialize a descriptor into a Parquet file stored under the dataset's duckdb directory."""
        dataset_root = dataset_directory(dataset_id)
        duckdb_dir = DatasetStorageService.dataset_duckdb_directory(dataset_id)

        if existing_relative_path:
            output_path = (dataset_root / existing_relative_path).resolve()
            output_path.parent.mkdir(parents=True, exist_ok=True)
            stored_name = output_path.name
            if output_path.exists():
                output_path.unlink()
        else:
            base_name = Path(descriptor.filename).stem or descriptor.alias or "file"
            stored_name = DatasetStorageService._deduplicate_filename(duckdb_dir, f"{base_name}.parquet")
            output_path = duckdb_dir / stored_name

        conn = duckdb.connect(database=":memory:", read_only=False)
        try:
            DuckDBService._prepare_connection(conn, [descriptor])
            reader_sql = DuckDBService._reader_sql(descriptor)
            escaped_output = str(output_path).replace("'", "''")
            conn.execute(
                f"COPY (SELECT * FROM {reader_sql}) TO '{escaped_output}' "
                "(FORMAT PARQUET, COMPRESSION 'ZSTD', OVERWRITE_OR_IGNORE TRUE)"
            )
            row_count = int(conn.execute(f"SELECT COUNT(*) FROM {reader_sql}").fetchone()[0])
        finally:
            conn.close()

        checksum = DataFrameFileService._compute_file_checksum(output_path)
        size = output_path.stat().st_size

        storage = StoredFileMetadata(
            dataset_id=dataset_id,
            original_filename=descriptor.filename,
            stored_filename=stored_name,
            absolute_path=output_path,
            relative_path=output_path.relative_to(dataset_root),
            size=size,
            checksum=checksum,
        )

        return ColumnarArtifact(storage=storage, format="parquet", row_count=row_count)

    @staticmethod
    async def _materialize_descriptor_to_parquet(
        dataset_id: str,
        descriptor: DuckDBFileDescriptor,
        existing_relative_path: str | None = None,
    ) -> ColumnarArtifact:
        """Async wrapper to materialize a descriptor into Parquet using a worker thread."""
        return await asyncio.to_thread(
            DataFrameFileService._materialize_descriptor_to_parquet_sync,
            dataset_id,
            descriptor,
            existing_relative_path,
        )

    @staticmethod
    def _count_rows_for_descriptor_sync(descriptor: DuckDBFileDescriptor) -> int:
        """Count rows for a descriptor using DuckDB synchronously."""
        conn = duckdb.connect(database=":memory:", read_only=False)
        try:
            DuckDBService._prepare_connection(conn, [descriptor])
            reader_sql = DuckDBService._reader_sql(descriptor)
            result = conn.execute(f"SELECT COUNT(*) FROM {reader_sql}").fetchone()
            return int(result[0]) if result else 0
        finally:
            conn.close()

    @staticmethod
    def _sheet_artifact_filename(file_record: File, sheet_name: str) -> str:
        """Generate a stable filename for an Excel sheet materialization."""
        base = DataFrameFileService._alias_from_filename(file_record.name).strip() or "sheet"
        sheet_component = re.sub(r"[^A-Za-z0-9]+", "_", sheet_name).strip("_") or "sheet"
        return f"{base}__{sheet_component}.parquet"

    @staticmethod
    def _convert_excel_sheet_to_parquet_sync(
        source_path: Path,
        sheet_name: str,
        target_path: Path,
    ) -> None:
        """Convert a single Excel sheet to Parquet using a CSV intermediary."""
        workbook = DataFrameFileService._open_excel_workbook(source_path)
        temp_csv: Path | None = None
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook {source_path}")

            sheet = workbook[sheet_name]
            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".csv", mode="w", encoding="utf-8", newline=""
            ) as tmp_file:
                temp_csv = Path(tmp_file.name)
                writer = csv.writer(tmp_file, lineterminator="\n")
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(["" if cell is None else cell for cell in row])
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        if temp_csv is None:
            raise ValueError(f"Failed to create temporary CSV for sheet '{sheet_name}'")

        target_path.parent.mkdir(parents=True, exist_ok=True)
        escaped_csv = str(temp_csv).replace("'", "''")
        escaped_target = str(target_path).replace("'", "''")

        conn = duckdb.connect(database=":memory:", read_only=False)
        try:
            copy_sql = (
                f"COPY (SELECT * FROM read_csv_auto('{escaped_csv}', HEADER=TRUE)) "
                f"TO '{escaped_target}' (FORMAT PARQUET, COMPRESSION 'ZSTD', OVERWRITE_OR_IGNORE TRUE)"
            )
            conn.execute(copy_sql)
        finally:
            conn.close()
            try:
                temp_csv.unlink()
            except FileNotFoundError:
                pass

    @staticmethod
    async def _ensure_excel_sheet_parquet(
        dataset,
        file_record: File,
        source_path: Path,
        sheet_name: str,
    ) -> Path:
        """Ensure a sheet-specific Parquet materialization exists for Excel fallback."""
        duckdb_dir = DatasetStorageService.dataset_duckdb_directory(str(dataset.id))
        artifact_name = DataFrameFileService._sheet_artifact_filename(file_record, sheet_name)
        target_path = (duckdb_dir / artifact_name).resolve()

        source_mtime = source_path.stat().st_mtime
        if target_path.exists() and target_path.stat().st_mtime >= source_mtime:
            return target_path

        lock = await DataFrameFileService._get_artifact_lock(file_record.id)
        async with lock:
            if target_path.exists() and target_path.stat().st_mtime >= source_mtime:
                return target_path
            await asyncio.to_thread(
                DataFrameFileService._convert_excel_sheet_to_parquet_sync,
                source_path,
                sheet_name,
                target_path,
            )
        return target_path

    @staticmethod
    async def _count_rows_for_descriptor(descriptor: DuckDBFileDescriptor) -> int:
        """Async helper to count rows for a descriptor."""
        return await asyncio.to_thread(
            DataFrameFileService._count_rows_for_descriptor_sync,
            descriptor,
        )

    @staticmethod
    async def save_file_to_db(
        session: AsyncSession,
        filename: str,
        dataset_id: str,
        file_type: str = "csv",
        file_content: bytes | None = None,
        storage_metadata: StoredFileMetadata | None = None,
        source_url: str | None = None,
        tenant_id: UUID | None = None,
    ) -> File:
        """
        Save uploaded file content to database.

        Args:
            session: Database session
            filename: Original filename
            dataset_id: Dataset ID this file belongs to
            file_type: File type (csv, excel, parquet, json)
            file_content: Optional in-memory content (legacy path)
            storage_metadata: Optional filesystem metadata (preferred path)
            source_url: Optional original source URL if downloaded from URL

        Returns:
            Created File model instance
        """
        try:
            safe_filename = re.sub(r"[^\w\-_\.]", "_", filename)

            if storage_metadata:
                stored_name = safe_filename
                size = storage_metadata.size
                storage_path = str(storage_metadata.relative_path)
                checksum = storage_metadata.checksum
                content = None
            else:
                if file_content is None:
                    raise ValueError("Either file_content or storage_metadata must be provided")
                stored_name = safe_filename
                size = len(file_content)
                storage_path = None
                checksum = hashlib.sha256(file_content).hexdigest()
                content = file_content

            file_repo = FileRepository(session)
            file_record = await file_repo.create(
                {
                    "name": stored_name,
                    "content": content,
                    "type": file_type,
                    "size": size,
                    "dataset_id": dataset_id,
                    "storage_path": storage_path,
                    "checksum": checksum,
                    "source_url": source_url,
                    "tenant_id": tenant_id,
                }
            )

            await session.commit()
            await session.refresh(file_record)

            logger.info(f"Saved {file_type} file to database: {safe_filename} (size: {size} bytes)")
            return file_record

        except Exception as e:
            logger.error(f"Error saving {file_type} file to database: {str(e)}")
            raise

    @staticmethod
    async def ensure_columnar_artifact(
        session: AsyncSession,
        dataset,
        file_record: File,
        storage_metadata: StoredFileMetadata | None = None,
    ) -> File | None:
        """
        Ensure a columnar (Parquet or equivalent) artifact exists for a file, updating the record in place.
        """
        lock = await DataFrameFileService._get_artifact_lock(file_record.id)
        # Prevent concurrent materialisation attempts for the same file which could race on filesystem writes.
        async with lock:
            return await DataFrameFileService._ensure_columnar_artifact_locked(
                session=session,
                dataset=dataset,
                file_record=file_record,
                storage_metadata=storage_metadata,
            )

    @staticmethod
    async def _ensure_columnar_artifact_locked(
        session: AsyncSession,
        dataset,
        file_record: File,
        storage_metadata: StoredFileMetadata | None = None,
    ) -> File | None:
        """
        Internal implementation of ensure_columnar_artifact guarded by a per-file lock.
        """
        file_type = (file_record.type or "").lower()
        supported_materialization = {"csv", "json"}
        already_columnar = {"parquet"}

        # Ensure dataset has a storage root established
        dataset_root = DataFrameFileService._dataset_root(dataset)

        source_path: Path | None = None
        if storage_metadata:
            source_path = storage_metadata.absolute_path
        elif file_record.storage_path:
            source_candidate = (dataset_root / file_record.storage_path).resolve()
            if source_candidate.exists():
                source_path = source_candidate
        if source_path is None:
            try:
                source_path = await DataFrameFileService._ensure_file_materialized(session, dataset, file_record)
            except Exception as exc:
                logger.warning(
                    "Unable to materialize file %s (%s) to disk for columnar conversion: %s",
                    file_record.id,
                    file_record.name,
                    exc,
                )
                return None

        alias = DataFrameFileService._alias_from_filename(file_record.name)

        optimized_candidate: Path | None = None
        if file_record.optimized_storage_path:
            candidate = (dataset_root / file_record.optimized_storage_path).resolve()
            if candidate.exists():
                optimized_candidate = candidate
            else:
                logger.warning(
                    "Optimized artifact missing for file %s (dataset %s). Will regenerate.",
                    file_record.id,
                    dataset.id,
                )

        # Columnar source files (Parquet) just need metadata cleanup.
        if file_type in already_columnar:
            dirty = False

            if optimized_candidate is None:
                optimized_candidate = source_path
                if file_record.storage_path and not file_record.optimized_storage_path:
                    file_record.optimized_storage_path = file_record.storage_path
                    dirty = True
                elif source_path is not None and not file_record.optimized_storage_path:
                    try:
                        relative = source_path.relative_to(dataset_root)
                        file_record.optimized_storage_path = str(relative)
                        dirty = True
                    except ValueError:
                        # Source lives outside dataset root, keep absolute path.
                        file_record.optimized_storage_path = str(source_path)
                        dirty = True

            if file_record.optimized_format != file_type:
                file_record.optimized_format = file_type
                dirty = True

            if file_record.optimized_checksum is None and file_record.checksum:
                file_record.optimized_checksum = file_record.checksum
                dirty = True

            if file_record.row_count is None and optimized_candidate and optimized_candidate.exists():
                descriptor = DuckDBFileDescriptor(
                    alias=alias,
                    path=optimized_candidate,
                    file_type=file_type,
                    filename=file_record.name,
                )
                try:
                    file_record.row_count = await DataFrameFileService._count_rows_for_descriptor(descriptor)
                    dirty = True
                except Exception as exc:
                    logger.warning(
                        "Failed to compute row count for %s: %s",
                        file_record.name,
                        exc,
                    )

            if dirty:
                session.add(file_record)
                await session.commit()
                await session.refresh(file_record)

            return file_record

        if file_type not in supported_materialization:
            return None

        # If an optimized Parquet snapshot already exists, reuse it.
        if optimized_candidate and optimized_candidate.exists():
            dirty = False
            if file_record.optimized_format != "parquet":
                file_record.optimized_format = "parquet"
                dirty = True
            if file_record.row_count is None:
                descriptor = DuckDBFileDescriptor(
                    alias=alias,
                    path=optimized_candidate,
                    file_type=file_record.optimized_format or "parquet",
                    filename=file_record.name,
                )
                try:
                    file_record.row_count = await DataFrameFileService._count_rows_for_descriptor(descriptor)
                    dirty = True
                except Exception as exc:
                    logger.warning(
                        "Failed to compute row count for %s optimized snapshot: %s",
                        file_record.name,
                        exc,
                    )
            if dirty:
                session.add(file_record)
                await session.commit()
                await session.refresh(file_record)
            return file_record

        descriptor = DuckDBFileDescriptor(
            alias=alias,
            path=source_path,
            file_type=file_type,
            filename=file_record.name,
        )

        try:
            artifact = await DataFrameFileService._materialize_descriptor_to_parquet(
                dataset_id=str(dataset.id),
                descriptor=descriptor,
                existing_relative_path=file_record.optimized_storage_path,
            )
        except Exception as exc:
            logger.warning(
                "Failed to create Parquet snapshot for %s (dataset %s): %s",
                file_record.name,
                dataset.id,
                exc,
            )
            return None

        file_record.optimized_storage_path = str(artifact.storage.relative_path)
        file_record.optimized_format = artifact.format
        file_record.optimized_checksum = artifact.storage.checksum
        file_record.row_count = artifact.row_count

        session.add(file_record)
        await session.commit()
        await session.refresh(file_record)

        logger.info(
            "Materialized columnar artifact for %s (%s rows) at %s",
            file_record.name,
            artifact.row_count,
            artifact.storage.relative_path,
        )

        return file_record

    @staticmethod
    async def get_file_content(session: AsyncSession, file_id: str) -> bytes:
        """
        Retrieve file content from database.

        Args:
            session: Database session
            file_id: File ID

        Returns:
            File content as bytes

        Raises:
            ValueError: If file not found
        """
        file_repo = FileRepository(session)
        file = await file_repo.get(file_id)
        if not file:
            raise ValueError(f"File {file_id} not found")

        logger.info(f"Retrieved file content for {file.name} ({file.size} bytes)")
        return file.content

    @staticmethod
    def detect_encoding(file_content: bytes) -> str:
        """
        Detect file encoding using chardet (for CSV files).

        Args:
            file_content: File content as bytes

        Returns:
            Detected encoding (defaults to 'utf-8' if detection fails)
        """
        try:
            if HAS_CHARDET:
                # Read first 100KB for detection
                sample = file_content[:100000]
                result = chardet.detect(sample)
                encoding = result.get("encoding", "utf-8")
                confidence = result.get("confidence", 0)

                logger.info(f"Detected encoding: {encoding} (confidence: {confidence:.2%})")

                # Fallback to common encodings if confidence is low
                if confidence < 0.7:
                    logger.warning(f"Low confidence encoding detection ({confidence:.2%}), trying common encodings")
                    for enc in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
                        try:
                            file_content.decode(enc)
                            logger.info(f"Successfully validated encoding: {enc}")
                            return enc
                        except (UnicodeDecodeError, UnicodeError):
                            continue

                return encoding or "utf-8"
            else:
                logger.warning("chardet not installed, falling back to common encodings")
                for enc in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
                    try:
                        file_content.decode(enc)
                        logger.info(f"Successfully validated encoding: {enc}")
                        return enc
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                return "utf-8"
        except Exception as e:
            logger.warning(f"Error detecting encoding: {str(e)}, using utf-8")
            return "utf-8"

    @staticmethod
    def get_excel_sheets(file_source: bytes | str | Path) -> list[str]:
        """
        Get list of all sheet names in an Excel file.

        Args:
            file_content: Excel file content as bytes

        Returns:
            List of sheet names
        """
        try:
            workbook = DataFrameFileService._open_excel_workbook(file_source)
            try:
                sheet_names = list(workbook.sheetnames)
                logger.info("Found %s sheets in Excel file: %s", len(sheet_names), sheet_names)
                return sheet_names
            finally:
                try:
                    workbook.close()
                except Exception:
                    pass
        except Exception as e:
            logger.error(f"Error reading Excel sheet names: {str(e)}")
            raise

    @staticmethod
    async def get_file_schema(
        file_content: bytes | None,
        filename: str,
        file_type: str = "csv",
        *,
        file_path: Path | str | None = None,
        sample_rows: int = 10,
        **_kwargs,
    ) -> dict[str, Any]:
        """Extract schema information for file datasets."""
        cleanup_path: Path | None = None
        try:
            if isinstance(file_content, (str, Path)) and file_path is None:
                file_path = Path(file_content)
                file_content = None

            file_type_lower = (file_type or "").lower()
            if not file_type_lower:
                raise ValueError("file_type is required for schema extraction")

            supported_types = set(DuckDBService._SQL_READERS.keys())
            if file_type_lower not in supported_types:
                supported_list = ", ".join(sorted(supported_types))
                raise ValueError(
                    f"DuckDB query support not implemented for file type '{file_type}'. "
                    f"Supported types: {supported_list}."
                )

            path_obj = Path(file_path) if file_path else None
            if path_obj is None:
                if file_content is None:
                    raise ValueError("file_content or file_path must be provided for schema extraction")
                temp_typ = "xlsx" if file_type_lower == "excel" else file_type_lower
                cleanup_path = DataFrameFileService._materialize_temp_file(file_content, temp_typ)
                path_obj = cleanup_path

            if file_type_lower == "excel":
                sheet_names = DataFrameFileService.get_excel_sheets(path_obj)
                if not sheet_names:
                    raise ValueError(f"Excel file '{filename}' contains no sheets")

                base_alias = DataFrameFileService._alias_from_filename(filename).strip() or "excel_table"
                alias_counts: dict[str, int] = {}
                multi_sheet = len(sheet_names) > 1
                descriptors: list[DuckDBFileDescriptor] = []

                for idx, sheet_name in enumerate(sheet_names):
                    sheet_alias = DataFrameFileService._alias_from_filename(sheet_name).strip() or f"sheet_{idx + 1}"
                    candidate_alias = base_alias if not multi_sheet else f"{base_alias}__{sheet_alias}"
                    if not candidate_alias:
                        candidate_alias = f"{base_alias}_sheet_{idx + 1}"

                    alias = DataFrameFileService._next_alias(alias_counts, candidate_alias)

                    descriptors.append(
                        DuckDBFileDescriptor(
                            alias=alias,
                            path=path_obj,
                            file_type=file_type_lower,
                            filename=filename,
                            sheet_name=sheet_name,
                            reader_options={"sheet": sheet_name},
                        )
                    )

                duckdb_schema = await DuckDBService.collect_schema(descriptors, sample_rows=sample_rows)
                schema_tables = duckdb_schema.get("schema", {})
                sample_by_table = duckdb_schema.get("sample_data", {})

                response: dict[str, Any] = {
                    "datasource_type": "excel",
                    "datasource_name": filename,
                    "filename": filename,
                    "schema": {},
                    "sample_data": sample_by_table,
                }

                for table_alias, table_info in schema_tables.items():
                    normalized = dict(table_info)
                    normalized.setdefault("filename", filename)
                    normalized.setdefault("file_type", "excel")
                    if "sheet_name" not in normalized:
                        descriptor = next((d for d in descriptors if d.alias == table_alias), None)
                        if descriptor and descriptor.sheet_name:
                            normalized["sheet_name"] = descriptor.sheet_name
                    response["schema"][table_alias] = normalized

                if schema_tables:
                    first_alias = next(iter(schema_tables))
                    first_info = schema_tables[first_alias]
                    response["columns"] = list(first_info.get("columns", []))
                    response["row_count"] = first_info.get("row_count", 0)
                else:
                    response["columns"] = []
                    response["row_count"] = 0

                logger.info(
                    "Generated Excel schema for %s with %s sheet(s)",
                    filename,
                    len(schema_tables),
                )
                return response

            descriptor = DuckDBFileDescriptor(
                alias=DataFrameFileService._alias_from_filename(filename),
                path=path_obj,
                file_type=file_type_lower,
                filename=filename,
            )
            schema = await DataFrameFileService._collect_duckdb_single_file_schema(
                descriptor,
                sample_rows=sample_rows,
            )
            logger.info(
                "Generated %s schema for %s: %s columns, %s rows (sample %s)",
                file_type_lower.upper(),
                filename,
                len(schema.get("columns", [])),
                schema.get("row_count", 0),
                len(schema.get("sample_data", [])),
            )
            return schema
        except Exception as e:
            logger.error(f"Error extracting {file_type} schema: {str(e)}")
            raise
        finally:
            if cleanup_path is not None:
                try:
                    os.unlink(cleanup_path)
                except FileNotFoundError:
                    pass
                except Exception as cleanup_err:
                    logger.warning("Failed to remove temporary file %s: %s", cleanup_path, cleanup_err)

    @staticmethod
    async def get_file_schema_multi(
        files: list[Any],
        session: AsyncSession | None = None,
        dataset: Any | None = None,
        sample_rows: int = 10,
        use_cache: bool = False,
        save_to_cache: bool = False,
    ) -> dict[str, Any]:
        """
        Generate unified schema for multiple files from database.

        Args:
            files: List of File model instances or file descriptors (dicts)
            session: Async session used to materialize file paths when needed
            dataset: Optional dataset object for File instances
            sample_rows: Number of rows to sample for preview
            use_cache: If True and dataset provided, try to use cached schema
            save_to_cache: If True and dataset provided, save generated schema to cache

        Returns:
            Unified schema dict in SQL-like multi-table format
        """
        try:
            # Try to use cached schema if requested and available
            if use_cache and dataset is not None:
                cached_schema = DataFrameFileService.get_cached_schema(dataset)
                if cached_schema:
                    logger.info(f"Using cached schema for dataset {dataset.id}")
                    return cached_schema
            if not files:
                return {
                    "datasource_type": "duckdb",
                    "datasource_name": "DuckDB File Dataset",
                    "schema": {},
                    "sample_data": {},
                }

            merged_schema = {
                "datasource_type": "duckdb",
                "datasource_name": "DuckDB File Dataset",
                "schema": {},
                "sample_data": {},
            }

            for item in files:
                if isinstance(item, File):
                    if session is None:
                        raise ValueError("session is required to generate schema from File models")
                    dataset_ref = dataset or getattr(item, "dataset", None)
                    if dataset_ref is None:
                        raise ValueError("dataset context required for File models")
                    file_path = await DataFrameFileService._ensure_file_materialized(session, dataset_ref, item)
                    alias = DataFrameFileService._alias_from_filename(item.name)
                    file_schema = await DataFrameFileService.get_file_schema(
                        None,
                        item.name,
                        item.type,
                        file_path=file_path,
                        sample_rows=sample_rows,
                    )
                else:
                    file_type = item.get("type")
                    raw_path = item.get("path") or item.get("file_path")
                    if not raw_path:
                        raise ValueError("File descriptor must include 'path' for schema generation")
                    path_obj = Path(raw_path)
                    file_name = item.get("name") or item.get("filename") or path_obj.name
                    alias = DataFrameFileService._alias_from_filename(file_name)
                    file_schema = await DataFrameFileService.get_file_schema(
                        None,
                        file_name,
                        file_type,
                        file_path=path_obj,
                        sample_rows=sample_rows,
                    )

                tables = file_schema.get("schema", {})
                sample_section = file_schema.get("sample_data", {})
                if tables:
                    for table_name, table_info in tables.items():
                        merged_schema["schema"][table_name] = table_info
                        sample = []
                        if isinstance(sample_section, dict):
                            sample = sample_section.get(table_name, table_info.get("sample_data", []))
                        elif isinstance(sample_section, list):
                            sample = sample_section
                        elif isinstance(table_info, dict):
                            sample = table_info.get("sample_data", [])
                        if sample:
                            merged_schema["sample_data"][table_name] = sample
                else:
                    merged_schema["schema"][alias] = {
                        "filename": file_schema.get("filename", alias),
                        "file_type": file_schema.get("datasource_type", "file"),
                        "columns": file_schema.get("columns", []),
                        "row_count": file_schema.get("row_count", 0),
                    }
                    samples = file_schema.get("sample_data", [])
                    if isinstance(samples, dict):
                        samples = samples.get(alias, [])
                    if samples:
                        merged_schema["sample_data"][alias] = samples

            logger.info("Generated multi-file schema for %s files", len(files))

            # Save to cache if requested
            if save_to_cache and dataset is not None and session is not None:
                try:
                    await DataFrameFileService.save_schema_cache(session, dataset, merged_schema)
                except Exception as cache_error:
                    # Log but don't fail the request if caching fails
                    logger.warning(f"Failed to cache schema for dataset {dataset.id}: {cache_error}")

            return merged_schema

        except Exception as e:
            logger.error(f"Error generating multi-file schema: {str(e)}")
            raise

    @staticmethod
    async def execute_duckdb_query(
        connection_obj: dict[str, Any],
        query: str,
        limit: int = 5,
        timeout: int = 30,
    ) -> dict[str, Any]:
        """
        Execute a DuckDB SQL query against a file dataset connection object.
        """
        from server.db.session import get_async_session

        limit = min(limit, 50) if limit else 5

        dataset_id = connection_obj.get("dataset_id")
        if not dataset_id:
            return {
                "success": False,
                "error": "No dataset_id found in connection object. File datasource not properly configured.",
            }

        async for session in get_async_session():
            try:
                return await DataFrameFileService.execute_duckdb_query_on_dataset(
                    session=session,
                    dataset_id=dataset_id,
                    query=query,
                    limit=limit,
                    timeout=timeout,
                )
            finally:
                await session.close()

    @staticmethod
    async def execute_duckdb_query_on_dataset(
        session: AsyncSession,
        dataset_id: str,
        query: str,
        limit: int = 5,
        timeout: int = 30,
    ) -> dict[str, Any]:
        """Execute DuckDB SQL query on dataset files."""
        try:
            normalized_limit: int | None
            if limit is None:
                normalized_limit = None
            elif limit <= 0:
                normalized_limit = None
            else:
                normalized_limit = min(limit, 50)
            dataset = await DatasetService.get_dataset(session, dataset_id)
            if not dataset or dataset.type != "file":
                raise ValueError(f"Dataset {dataset_id} not found or not a file dataset")

            for file in list(dataset.files or []):
                try:
                    await DataFrameFileService.ensure_columnar_artifact(
                        session=session,
                        dataset=dataset,
                        file_record=file,
                    )
                except Exception as exc:
                    logger.warning(
                        "Columnar materialization skipped for file %s (dataset %s): %s",
                        file.id,
                        dataset_id,
                        exc,
                    )

            # Reload dataset to capture any updated file metadata (optimized paths, row counts)
            dataset = await DatasetService.get_dataset(session, dataset_id)
            if not dataset or dataset.type != "file":
                raise ValueError(f"Dataset {dataset_id} became unavailable during query preparation")

            duckdb_dir = DatasetStorageService.dataset_duckdb_directory(str(dataset.id))
            db_path = Path(dataset.duckdb_path) if dataset.duckdb_path else duckdb_dir / "dataset.duckdb"
            file_types = {str(file.type or "").lower() for file in list(dataset.files or [])}
            use_existing_catalog = (
                bool(dataset.duckdb_path)
                and db_path.exists()
                and db_path.suffix.lower() == ".duckdb"
                and bool(file_types)
                and file_types.issubset({"duckdb"})
            )
            descriptors = [] if use_existing_catalog else await DataFrameFileService._build_duckdb_descriptors(session, dataset)

            if not dataset.duckdb_path:
                dataset.duckdb_path = str(db_path)
                session.add(dataset)
                try:
                    await session.flush()
                except Exception:
                    await session.rollback()
                    raise

            return await DuckDBService.execute_sql(
                descriptors,
                query,
                limit=normalized_limit,
                database_path=db_path,
                timeout=timeout,
            )
        except ValueError as e:
            logger.error(f"Validation error executing DuckDB query: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "query": query,
            }
        except Exception as e:
            logger.error(f"Error executing DuckDB query: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "query": query,
            }

    FILE_TYPE_ABBREV = {
        "varchar": "str",
        "character varying": "str",
        "text": "str",
        "integer": "int",
        "bigint": "int",
        "smallint": "int",
        "timestamp": "ts",
        "timestamp without time zone": "ts",
        "decimal": "dec",
        "numeric": "dec",
        "boolean": "bool",
        "double precision": "float",
        "double": "float",
        "real": "float",
        "date": "date",
        "time": "time",
    }

    @staticmethod
    def _abbreviate_file_type(full_type: str) -> str:
        normalized = full_type.lower().strip()
        base_type = normalized.split("(")[0].strip()
        return DataFrameFileService.FILE_TYPE_ABBREV.get(base_type, base_type)

    @staticmethod
    def _format_file_column_compact(col: dict[str, Any]) -> str:
        col_name = col.get("name", "unknown")
        col_type = DataFrameFileService._abbreviate_file_type(col.get("type", "unknown"))
        nullable = col.get("nullable", True)
        not_null_marker = "!" if not nullable else ""

        annotation = col.get("annotation", "")
        if annotation:
            return f"{col_name}({col_type}{not_null_marker}) [{annotation}]"

        return f"{col_name}({col_type}{not_null_marker})"

    @staticmethod
    def format_file_schema_for_prompt(schema_data: dict[str, Any]) -> str:
        """Format file schema into a compact, token-efficient string for AI prompts."""
        lines: list[str] = []

        schema_tables = schema_data.get("schema", {})
        datasource_type = (schema_data.get("datasource_type") or "duckdb").lower()

        if datasource_type == "excel":
            type_label = "Excel"
        else:
            type_label = "DuckDB"

        if len(schema_tables) > 1:
            filenames = {table_info.get("filename") for table_info in schema_tables.values()}
            is_multi_sheet_excel = len(filenames) == 1 and datasource_type == "excel"

            if is_multi_sheet_excel:
                filename = next(iter(filenames)) or "unknown"
                lines.append(f"[{type_label}:{filename} ({len(schema_tables)} sheets)]")
            else:
                lines.append(f"[{type_label}:{len(schema_tables)} files]")

            for table_name, table_info in schema_tables.items():
                row_count = table_info.get("row_count", 0)
                cols = []
                for col in table_info.get("columns", []):
                    cols.append(DataFrameFileService._format_file_column_compact(col))

                desc = table_info.get("description", "")
                line = f"{table_name} ({row_count} rows): {', '.join(cols)}"
                if desc:
                    line += f" // {desc}"
                lines.append(line)
        else:
            filename = schema_data.get("filename", "unknown")
            row_count = schema_data.get("row_count", 0)
            lines.append(f"[{type_label}:{filename} ({row_count} rows)]")

            columns = schema_data.get("columns", [])
            if columns:
                cols = []
                for col in columns:
                    cols.append(DataFrameFileService._format_file_column_compact(col))

                table_name = Path(filename).stem if filename else "data"
                desc = schema_data.get("description", "")
                line = f"{table_name}: {', '.join(cols)}"
                if desc:
                    line += f" // {desc}"
                lines.append(line)

        return "\n".join(lines)
