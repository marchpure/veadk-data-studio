"""Build multi-sheet xlsx workbook from parsed Slack tables."""

from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook

EXCEL_CELL_MAX = 32767
SHEET_NAME_MAX = 31
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sanitize_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("", name).strip() or "Sheet"
    cleaned = cleaned[:SHEET_NAME_MAX]

    candidate = cleaned
    suffix = 2
    while candidate in existing:
        marker = f"_{suffix}"
        candidate = (cleaned[: SHEET_NAME_MAX - len(marker)]) + marker
        suffix += 1

    existing.add(candidate)
    return candidate


def build_xlsx_from_tables(tables: list[list[list[str]]]) -> bytes:
    """
    Build an xlsx workbook with one sheet per table.

    Args:
        tables: List of tables. Each table is a list of rows (first row = header).

    Returns:
        Bytes of the .xlsx file.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names: set[str] = set()

    for idx, table in enumerate(tables):
        if not table:
            continue

        sheet_name = _sanitize_sheet_name(f"Query {idx + 1}", used_names)
        worksheet = workbook.create_sheet(title=sheet_name)

        for row in table:
            worksheet.append([(cell[:EXCEL_CELL_MAX] if isinstance(cell, str) else cell) for cell in row])

    if not workbook.sheetnames:
        workbook.create_sheet(title="Empty")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
