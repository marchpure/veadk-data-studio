from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import re
import secrets
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Protocol
from urllib.parse import unquote, urlencode, urlparse
from uuid import UUID

import httpx
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from server.models.source_connections import SourceConnection
from server.models.source_resources import SourceResource
from server.services.crypto_service import CryptoService
from server.services.settings import SettingsService
from server.services.source_redaction import redact_sensitive_json, source_ref
from server.utils.custom_logger import get_logger

logger = get_logger(__name__)

FEISHU_CONFIG_KEY = "source_connector_feishu_config"
FEISHU_OAUTH_STATE_TTL_SECONDS = 600
FEISHU_REFRESH_SKEW_SECONDS = 300

REQUIRED_FEISHU_SCOPES = [
    "drive:drive:readonly",
    "docs:doc:readonly",
    "wiki:wiki:readonly",
    "sheets:spreadsheet:readonly",
    "bitable:app:readonly",
]


class ConnectorError(Exception):
    def __init__(self, message: str, *, code: str = "connector_error", permanent: bool = False):
        super().__init__(message)
        self.code = code
        self.permanent = permanent


@dataclass(frozen=True)
class ResourceListInput:
    tenant_id: UUID
    connection: SourceConnection
    scope: str = "recent"
    parent_token: str | None = None
    resource_type: str | None = None
    query: str | None = None
    page_token: str | None = None
    page_size: int = 50
    already_added_external_ids: frozenset[str] = frozenset()


@dataclass(frozen=True)
class ResourcePickerItem:
    external_id: str
    resource_type: str
    name: str
    parent_external_id: str | None = None
    source_url: str | None = None
    has_children: bool = False
    is_folder: bool = False
    already_added: bool = False
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_payload(self) -> dict[str, Any]:
        return {
            "external_id": self.external_id,
            "resource_type": self.resource_type,
            "name": self.name,
            "parent_external_id": self.parent_external_id,
            "source_url": self.source_url,
            "has_children": self.has_children,
            "is_folder": self.is_folder,
            "already_added": self.already_added,
            "metadata": self.metadata,
        }


@dataclass(frozen=True)
class ResourceListResult:
    items: list[ResourcePickerItem]
    next_page_token: str | None = None


@dataclass(frozen=True)
class CapturedSnapshot:
    raw_bytes: bytes
    content_text: str
    external_revision: str | None
    metadata: dict[str, Any]
    provider: str
    parser_version: str
    raw_storage_uri: str


class SourceConnectorAdapter(Protocol):
    provider: str

    async def test_connection(self, credentials: dict[str, Any]) -> dict[str, Any]:
        ...

    async def list_resources(
        self,
        *,
        session: AsyncSession,
        input: ResourceListInput,
    ) -> ResourceListResult:
        ...

    async def sync_resource(
        self,
        *,
        session: AsyncSession,
        connection: SourceConnection,
        resource: SourceResource,
    ) -> CapturedSnapshot:
        ...


def mask_secret(value: str | None) -> str | None:
    if not value:
        return value
    if len(value) <= 8:
        return "****"
    return f"{value[:4]}…{value[-4:]}"


def redact_credentials(credentials: dict[str, Any]) -> dict[str, Any]:
    secret_keys = {"secret", "token", "access_key", "refresh", "password", "sk"}
    redacted: dict[str, Any] = {}
    for key, value in credentials.items():
        if any(term in key.lower() for term in secret_keys):
            redacted[key] = mask_secret(str(value)) if value is not None else None
        else:
            redacted[key] = value
    return redacted


class FeishuAdminConfigService:
    @staticmethod
    async def save_config(
        *,
        session: AsyncSession,
        app_id: str,
        app_secret: str,
        redirect_uri: str,
        scopes: list[str],
    ) -> None:
        encrypted = await CryptoService.encrypt_config(
            {
                "app_id": app_id.strip(),
                "app_secret": app_secret.strip(),
                "redirect_uri": redirect_uri.strip(),
                "scopes": scopes or REQUIRED_FEISHU_SCOPES,
            },
            session,
        )
        await SettingsService.upsert_setting(
            session,
            FEISHU_CONFIG_KEY,
            encrypted,
            description="Encrypted Feishu connector OAuth configuration",
            is_encrypted=True,
        )

    @staticmethod
    async def load_config(*, session: AsyncSession) -> dict[str, Any] | None:
        setting = await SettingsService.get_setting_by_key(session, FEISHU_CONFIG_KEY)
        if not setting or not setting.setting_value:
            return None
        return await CryptoService.decrypt_config(setting.setting_value, session)

    @staticmethod
    async def status(*, session: AsyncSession) -> dict[str, Any]:
        config = await FeishuAdminConfigService.load_config(session=session)
        if not config:
            return {
                "configured": False,
                "required_scopes": REQUIRED_FEISHU_SCOPES,
                "missing_scopes": REQUIRED_FEISHU_SCOPES,
            }
        scopes = config.get("scopes") or []
        missing = [scope for scope in REQUIRED_FEISHU_SCOPES if scope not in scopes]
        return {
            "configured": bool(config.get("app_id") and config.get("app_secret") and config.get("redirect_uri")),
            "app_id": config.get("app_id"),
            "redirect_uri": config.get("redirect_uri"),
            "scopes": scopes,
            "required_scopes": REQUIRED_FEISHU_SCOPES,
            "missing_scopes": missing,
        }


class FeishuOAuthStateStore:
    _states: dict[str, dict[str, Any]] = {}

    @classmethod
    def create(cls, *, tenant_id: UUID, user_id: UUID, redirect_uri: str) -> str:
        state = secrets.token_urlsafe(32)
        cls._states[state] = {
            "tenant_id": str(tenant_id),
            "user_id": str(user_id),
            "redirect_uri": redirect_uri,
            "created_at": datetime.utcnow(),
        }
        return state

    @classmethod
    def pop(cls, state: str) -> dict[str, Any] | None:
        value = cls._states.pop(state, None)
        if value is None:
            return None
        created_at = value.get("created_at")
        if not isinstance(created_at, datetime):
            return None
        if datetime.utcnow() - created_at > timedelta(seconds=FEISHU_OAUTH_STATE_TTL_SECONDS):
            return None
        return value


class FeishuConnectorAdapter:
    provider = "feishu"
    base_url = "https://open.feishu.cn"
    url_path_resource_types = {
        "doc": "feishu_doc",
        "docs": "feishu_doc",
        "docx": "feishu_doc",
        "wiki": "feishu_wiki",
        "sheets": "feishu_sheet",
        "base": "feishu_base",
        "bitable": "feishu_base",
    }

    async def create_authorization_url(
        self,
        *,
        session: AsyncSession,
        tenant_id: UUID,
        user_id: UUID,
    ) -> tuple[str, str]:
        config = await FeishuAdminConfigService.load_config(session=session)
        if not config:
            raise ConnectorError("Feishu application is not configured", code="admin_config_required", permanent=True)
        redirect_uri = str(config.get("redirect_uri") or "").strip()
        if not redirect_uri:
            raise ConnectorError("Feishu OAuth redirect URI is not configured", code="admin_config_required", permanent=True)
        state = FeishuOAuthStateStore.create(tenant_id=tenant_id, user_id=user_id, redirect_uri=redirect_uri)
        params = {
            "app_id": config["app_id"],
            "redirect_uri": redirect_uri,
            "state": state,
        }
        return f"{self.base_url}/open-apis/authen/v1/authorize?{urlencode(params)}", state

    async def complete_oauth_callback(
        self,
        *,
        session: AsyncSession,
        code: str,
        state: str,
    ) -> SourceConnection:
        state_payload = FeishuOAuthStateStore.pop(state)
        if state_payload is None:
            raise ConnectorError("Invalid or expired Feishu OAuth state", code="invalid_state", permanent=True)
        config = await FeishuAdminConfigService.load_config(session=session)
        if not config:
            raise ConnectorError("Feishu application is not configured", code="admin_config_required", permanent=True)

        token = await self._exchange_code(config=config, code=code)
        user_info = await self._get_user_info(token["access_token"])
        tenant_id = UUID(state_payload["tenant_id"])
        user_id = UUID(state_payload["user_id"])
        expires_at = datetime.utcnow() + timedelta(seconds=max(0, int(token.get("expires_in") or 0)))
        external_account_id = (
            user_info.get("open_id")
            or user_info.get("union_id")
            or user_info.get("user_id")
            or token.get("open_id")
            or token.get("union_id")
        )
        display_name = user_info.get("name") or user_info.get("en_name") or "Feishu"
        encrypted = await CryptoService.encrypt_config(
            {
                "access_token": token["access_token"],
                "refresh_token": token.get("refresh_token"),
                "scope": token.get("scope") or config.get("scopes") or [],
                "token_type": token.get("token_type"),
            },
            session,
        )

        existing = None
        if external_account_id:
            existing = await session.scalar(
                select(SourceConnection).where(
                    SourceConnection.tenant_id == tenant_id,
                    SourceConnection.provider == self.provider,
                    SourceConnection.created_by == user_id,
                    SourceConnection.external_account_id == external_account_id,
                )
            )
        if existing:
            connection = existing
            connection.encrypted_credentials = encrypted
            connection.display_name = display_name
            connection.status = "connected"
            connection.token_expires_at = expires_at
            connection.capabilities_json = {"scopes": token.get("scope") or config.get("scopes") or []}
        else:
            connection = SourceConnection(
                tenant_id=tenant_id,
                provider=self.provider,
                auth_mode="oauth",
                encrypted_credentials=encrypted,
                external_account_id=external_account_id,
                display_name=display_name,
                status="connected",
                capabilities_json={"scopes": token.get("scope") or config.get("scopes") or []},
                token_expires_at=expires_at,
                created_by=user_id,
            )
            session.add(connection)
        await session.commit()
        await session.refresh(connection)
        return connection

    async def test_connection(self, credentials: dict[str, Any]) -> dict[str, Any]:
        if not credentials.get("access_token"):
            raise ConnectorError("Feishu access token is missing", code="missing_token", permanent=True)
        return await self._get_user_info(str(credentials["access_token"]))

    async def ensure_access_token(self, *, session: AsyncSession, connection: SourceConnection) -> str:
        credentials = await CryptoService.decrypt_config(connection.encrypted_credentials, session)
        access_token = credentials.get("access_token")
        refresh_token = credentials.get("refresh_token")
        expires_at = connection.token_expires_at
        if access_token and expires_at and expires_at > datetime.utcnow() + timedelta(seconds=FEISHU_REFRESH_SKEW_SECONDS):
            return str(access_token)
        if not refresh_token:
            connection.status = "reauthorization_required"
            await session.commit()
            raise ConnectorError("Feishu refresh token is missing", code="reauthorization_required", permanent=True)
        config = await FeishuAdminConfigService.load_config(session=session)
        if not config:
            raise ConnectorError("Feishu application is not configured", code="admin_config_required", permanent=True)
        try:
            refreshed = await self._refresh_token(config=config, refresh_token=str(refresh_token))
        except ConnectorError as error:
            if error.permanent:
                connection.status = "reauthorization_required"
                await session.commit()
            raise
        credentials.update(
            {
                "access_token": refreshed["access_token"],
                "refresh_token": refreshed.get("refresh_token") or refresh_token,
                "scope": refreshed.get("scope") or credentials.get("scope"),
            }
        )
        connection.encrypted_credentials = await CryptoService.encrypt_config(credentials, session)
        connection.token_expires_at = datetime.utcnow() + timedelta(seconds=max(0, int(refreshed.get("expires_in") or 0)))
        connection.status = "connected"
        await session.commit()
        return str(refreshed["access_token"])

    async def list_resources(
        self,
        *,
        session: AsyncSession,
        input: ResourceListInput,
    ) -> ResourceListResult:
        access_token = await self.ensure_access_token(session=session, connection=input.connection)
        if input.scope == "wiki":
            return await self._list_wiki_resources(access_token=access_token, input=input)
        if input.query:
            return await self._search_drive_resources(access_token=access_token, input=input)
        return await self._list_drive_resources(access_token=access_token, input=input)

    async def locate_resource_from_url(
        self,
        *,
        access_token: str,
        url: str,
        already_added: frozenset[str],
    ) -> ResourcePickerItem:
        parsed = self.parse_resource_url(url)
        resource_type = parsed["resource_type"]
        token = parsed["external_id"]
        if resource_type == "feishu_doc":
            body = await self._request_json("GET", f"/open-apis/docx/v1/documents/{token}", access_token=access_token)
            document = ((body.get("data") or {}).get("document") or body.get("data") or {})
            return ResourcePickerItem(
                external_id=token,
                resource_type="feishu_doc",
                name=document.get("title") or document.get("name") or token,
                source_url=url,
                already_added=token in already_added,
                metadata={"type": "docx", "token": token, "document": document, "located_from_url": url},
            )
        if resource_type == "feishu_wiki":
            resolved = await self._resolve_wiki(access_token=access_token, wiki_token=token)
            return ResourcePickerItem(
                external_id=token,
                resource_type="feishu_wiki",
                name=resolved.get("title") or token,
                source_url=url,
                has_children=bool((resolved.get("node") or {}).get("has_child")),
                already_added=token in already_added,
                metadata={
                    "type": "wiki",
                    "node_token": token,
                    "obj_token": resolved.get("obj_token"),
                    "obj_type": (resolved.get("node") or {}).get("obj_type"),
                    "resolved_wiki": resolved,
                    "located_from_url": url,
                },
            )
        if resource_type == "feishu_sheet":
            body = await self._request_json("GET", f"/open-apis/sheets/v3/spreadsheets/{token}", access_token=access_token)
            spreadsheet = ((body.get("data") or {}).get("spreadsheet") or body.get("data") or {})
            return ResourcePickerItem(
                external_id=token,
                resource_type="feishu_sheet",
                name=spreadsheet.get("title") or spreadsheet.get("name") or token,
                source_url=url,
                already_added=token in already_added,
                metadata={"type": "sheet", "spreadsheet_token": token, "spreadsheet": spreadsheet, "located_from_url": url},
            )
        if resource_type == "feishu_base":
            body = await self._request_json("GET", f"/open-apis/bitable/v1/apps/{token}/tables", access_token=access_token)
            tables = ((body.get("data") or {}).get("items") or [])
            return ResourcePickerItem(
                external_id=token,
                resource_type="feishu_base",
                name=token,
                source_url=url,
                already_added=token in already_added,
                metadata={"type": "bitable", "app_token": token, "tables": tables, "located_from_url": url},
            )
        raise ConnectorError(f"Unsupported Feishu resource type: {resource_type}", code="unsupported_resource", permanent=True)

    async def sync_resource(
        self,
        *,
        session: AsyncSession,
        connection: SourceConnection,
        resource: SourceResource,
    ) -> CapturedSnapshot:
        access_token = await self.ensure_access_token(session=session, connection=connection)
        metadata: dict[str, Any] = dict(resource.selection_config_json or {})
        if resource.resource_type == "feishu_wiki":
            resolved = await self._resolve_wiki(access_token=access_token, wiki_token=resource.external_id or "")
            metadata["resolved_wiki"] = resolved
            actual_type = resolved.get("resource_type") or "feishu_doc"
            external_id = resolved.get("obj_token") or resource.external_id
        else:
            actual_type = resource.resource_type
            external_id = resource.external_id

        if actual_type == "feishu_doc":
            content, raw, revision = await self._fetch_docx(access_token=access_token, document_id=str(external_id))
        elif actual_type == "feishu_sheet":
            content, raw, revision = await self._fetch_sheet(
                access_token=access_token,
                spreadsheet_token=str(external_id),
                selection=metadata,
            )
        elif actual_type == "feishu_base":
            content, raw, revision = await self._fetch_bitable(
                access_token=access_token,
                app_token=str(external_id),
                selection=metadata,
            )
        else:
            raise ConnectorError(f"Unsupported Feishu resource type: {actual_type}", code="unsupported_resource", permanent=True)

        raw_bytes = json.dumps(raw, ensure_ascii=False, sort_keys=True).encode("utf-8")
        return CapturedSnapshot(
            raw_bytes=raw_bytes,
            content_text=content,
            external_revision=revision,
            metadata={
                **redact_sensitive_json(metadata, resource_type=resource.resource_type),
                "provider": self.provider,
                "resource_type": resource.resource_type,
                "actual_resource_type": actual_type,
                "external_id_ref": source_ref(f"{resource.resource_type}_external", resource.external_id),
                "source_url_ref": source_ref("url", resource.source_url),
            },
            provider="byaan-native",
            parser_version="feishu-openapi-v1",
            raw_storage_uri=f"feishu://{resource.resource_type}/{source_ref(resource.resource_type, resource.external_id)}",
        )

    async def _exchange_code(self, *, config: dict[str, Any], code: str) -> dict[str, Any]:
        payload = {
            "grant_type": "authorization_code",
            "code": code,
        }
        return await self._post_authen("/open-apis/authen/v1/access_token", config=config, payload=payload)

    async def _refresh_token(self, *, config: dict[str, Any], refresh_token: str) -> dict[str, Any]:
        payload = {
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
        }
        return await self._post_authen("/open-apis/authen/v1/refresh_access_token", config=config, payload=payload)

    async def _post_authen(self, path: str, *, config: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
        data = {
            "app_id": config["app_id"],
            "app_secret": config["app_secret"],
            **payload,
        }
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(f"{self.base_url}{path}", json=data)
        if response.status_code >= 400:
            raise ConnectorError(f"Feishu OAuth request failed with HTTP {response.status_code}", code="oauth_http_error")
        body = response.json()
        if body.get("code", 0) not in (0, None):
            code = body.get("code")
            msg = body.get("msg") or body.get("message") or "Feishu OAuth failed"
            raise ConnectorError(
                f"Feishu OAuth failed: code={code}, msg={msg}",
                code="reauthorization_required" if "refresh" in str(msg).lower() else "oauth_error",
                permanent=code in {20001, 20002, 20003},
            )
        token = body.get("data") or body
        if not token.get("access_token"):
            raise ConnectorError("Feishu OAuth response missing access_token", code="oauth_response_invalid")
        return token

    async def _get_user_info(self, access_token: str) -> dict[str, Any]:
        body = await self._request_json("GET", "/open-apis/authen/v1/user_info", access_token=access_token)
        return body.get("data") or body

    async def _request_json(
        self,
        method: str,
        path: str,
        *,
        access_token: str,
        params: dict[str, Any] | None = None,
        json_body: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.request(
                method,
                f"{self.base_url}{path}",
                headers={"Authorization": f"Bearer {access_token}"},
                params={key: value for key, value in (params or {}).items() if value not in (None, "")},
                json=json_body,
            )
        if response.status_code >= 400:
            if response.status_code in {401, 403}:
                raise ConnectorError(
                    f"Feishu permission denied with HTTP {response.status_code}",
                    code="permission_lost",
                    permanent=True,
                )
            if response.status_code == 404:
                raise ConnectorError(
                    "Feishu resource unavailable with HTTP 404",
                    code="source_unavailable",
                    permanent=True,
                )
            if response.status_code == 429:
                raise ConnectorError("Feishu rate limit exceeded", code="rate_limited")
            if response.status_code in {408, 504}:
                raise ConnectorError(f"Feishu API timeout with HTTP {response.status_code}", code="source_timeout")
            raise ConnectorError(f"Feishu API request failed with HTTP {response.status_code}", code="feishu_http_error")
        body = response.json()
        if body.get("code", 0) not in (0, None):
            code = body.get("code")
            msg = body.get("msg") or body.get("message") or "Feishu API failed"
            text = f"{code} {msg}".lower()
            if any(term in text for term in ("permission", "forbidden", "unauthorized")):
                raise ConnectorError(f"Feishu permission denied: code={code}, msg={msg}", code="permission_lost", permanent=True)
            if any(term in text for term in ("not found", "not exist", "deleted")):
                raise ConnectorError(f"Feishu resource unavailable: code={code}, msg={msg}", code="source_unavailable", permanent=True)
            if any(term in text for term in ("rate limit", "too many request", "too many requests", "qps")):
                raise ConnectorError(f"Feishu rate limited: code={code}, msg={msg}", code="rate_limited")
            if any(term in text for term in ("timeout", "timed out")):
                raise ConnectorError(f"Feishu API timeout: code={code}, msg={msg}", code="source_timeout")
            raise ConnectorError(f"Feishu API failed: code={code}, msg={msg}", code="feishu_api_error")
        return body

    async def _list_drive_resources(self, *, access_token: str, input: ResourceListInput) -> ResourceListResult:
        body = await self._request_json(
            "GET",
            "/open-apis/drive/v1/files",
            access_token=access_token,
            params={
                "folder_token": input.parent_token or "",
                "page_token": input.page_token,
                "page_size": input.page_size,
                "order_by": "EditedTime",
                "direction": "DESC",
            },
        )
        data = body.get("data") or {}
        items = [self._drive_item_to_picker(item, input.already_added_external_ids) for item in data.get("files", []) or data.get("items", [])]
        return ResourceListResult(items=[item for item in items if self._matches_type(item, input.resource_type)], next_page_token=data.get("next_page_token") or data.get("page_token"))

    async def _search_drive_resources(self, *, access_token: str, input: ResourceListInput) -> ResourceListResult:
        body = await self._request_json(
            "POST",
            "/open-apis/drive/v1/files/search",
            access_token=access_token,
            json_body={
                "search_key": input.query or "",
                "page_token": input.page_token,
                "page_size": input.page_size,
            },
        )
        data = body.get("data") or {}
        items = [self._drive_item_to_picker(item, input.already_added_external_ids) for item in data.get("files", []) or data.get("items", [])]
        return ResourceListResult(items=[item for item in items if self._matches_type(item, input.resource_type)], next_page_token=data.get("next_page_token") or data.get("page_token"))

    async def _list_wiki_resources(self, *, access_token: str, input: ResourceListInput) -> ResourceListResult:
        path = "/open-apis/wiki/v2/spaces"
        params = {"page_token": input.page_token, "page_size": input.page_size}
        if input.parent_token:
            path = f"/open-apis/wiki/v2/spaces/{input.parent_token}/nodes"
        body = await self._request_json("GET", path, access_token=access_token, params=params)
        data = body.get("data") or {}
        raw_items = data.get("items") or data.get("nodes") or []
        items = [self._wiki_item_to_picker(item, input.already_added_external_ids) for item in raw_items]
        return ResourceListResult(items=items, next_page_token=data.get("next_page_token") or data.get("page_token"))

    def _drive_item_to_picker(self, item: dict[str, Any], already_added: frozenset[str]) -> ResourcePickerItem:
        file_type = str(item.get("type") or item.get("file_type") or "").lower()
        token = str(item.get("token") or item.get("file_token") or item.get("document_id") or item.get("spreadsheet_token") or "")
        resource_type = {
            "docx": "feishu_doc",
            "doc": "feishu_doc",
            "sheet": "feishu_sheet",
            "sheets": "feishu_sheet",
            "bitable": "feishu_base",
            "base": "feishu_base",
            "folder": "feishu_folder",
        }.get(file_type, f"feishu_{file_type}" if file_type else "feishu_doc")
        return ResourcePickerItem(
            external_id=token,
            resource_type=resource_type,
            name=item.get("name") or item.get("title") or token,
            parent_external_id=item.get("parent_token"),
            source_url=item.get("url"),
            has_children=resource_type == "feishu_folder",
            is_folder=resource_type == "feishu_folder",
            already_added=token in already_added,
            metadata=item,
        )

    def _wiki_item_to_picker(self, item: dict[str, Any], already_added: frozenset[str]) -> ResourcePickerItem:
        token = str(item.get("node_token") or item.get("token") or item.get("space_id") or "")
        return ResourcePickerItem(
            external_id=token,
            resource_type="feishu_wiki",
            name=item.get("title") or item.get("name") or token,
            parent_external_id=item.get("parent_node_token"),
            source_url=item.get("url"),
            has_children=bool(item.get("has_child") or item.get("node_type") == "origin"),
            is_folder=bool(item.get("node_type") in {"origin", "space"}),
            already_added=token in already_added,
            metadata=item,
        )

    def _matches_type(self, item: ResourcePickerItem, resource_type: str | None) -> bool:
        if not resource_type:
            return True
        return item.resource_type == resource_type

    def parse_resource_url(self, url: str) -> dict[str, str]:
        parsed = urlparse(url.strip())
        path_parts = [unquote(part) for part in parsed.path.split("/") if part]
        for index, part in enumerate(path_parts):
            resource_type = self.url_path_resource_types.get(part.lower())
            if resource_type and index + 1 < len(path_parts):
                external_id = path_parts[index + 1].strip()
                if external_id:
                    return {"resource_type": resource_type, "external_id": external_id}
        raise ConnectorError("Unsupported Feishu link. Paste a Docx, Wiki, Sheet, or Base URL.", code="unsupported_feishu_link", permanent=True)

    async def _resolve_wiki(self, *, access_token: str, wiki_token: str) -> dict[str, Any]:
        body = await self._request_json("GET", "/open-apis/wiki/v2/spaces/get_node", access_token=access_token, params={"token": wiki_token})
        node = ((body.get("data") or {}).get("node") or {})
        obj_type = node.get("obj_type") or ""
        resource_type = {"doc": "feishu_doc", "docx": "feishu_doc", "sheet": "feishu_sheet", "bitable": "feishu_base"}.get(obj_type, f"feishu_{obj_type}")
        return {"resource_type": resource_type, "obj_token": node.get("obj_token"), "title": node.get("title"), "node": node}

    async def _fetch_docx(self, *, access_token: str, document_id: str) -> tuple[str, dict[str, Any], str | None]:
        blocks: list[dict[str, Any]] = []
        page_token = None
        while True:
            body = await self._request_json(
                "GET",
                f"/open-apis/docx/v1/documents/{document_id}/blocks",
                access_token=access_token,
                params={"page_token": page_token, "page_size": 500, "document_revision_id": -1},
            )
            data = body.get("data") or {}
            blocks.extend(data.get("items") or [])
            if not data.get("has_more"):
                break
            page_token = data.get("page_token")
        title = "Feishu Doc"
        lines: list[str] = []
        for block in blocks:
            text = self._extract_feishu_text(block)
            if text:
                lines.append(text)
            if not lines and (block.get("page") or {}).get("elements"):
                title = self._extract_feishu_text(block) or title
        return "\n\n".join(lines), {"blocks": blocks}, self._stable_revision(blocks)

    async def _fetch_sheet(
        self,
        *,
        access_token: str,
        spreadsheet_token: str,
        selection: dict[str, Any],
    ) -> tuple[str, dict[str, Any], str | None]:
        meta = await self._request_json("GET", f"/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}", access_token=access_token)
        sheets_body = await self._request_json("GET", f"/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/query", access_token=access_token)
        sheets = ((sheets_body.get("data") or {}).get("sheets") or [])
        selected = {item.get("sheet_id") for item in (selection.get("subresources") or []) if item.get("sheet_id")}
        lines = [f"# {(((meta.get('data') or {}).get('spreadsheet') or {}).get('title') or 'Spreadsheet')}"]
        raw = {"metadata": meta.get("data"), "sheets": []}
        for sheet in sheets:
            sheet_id = sheet.get("sheet_id")
            if selected and sheet_id not in selected:
                continue
            title = sheet.get("title") or sheet_id
            grid = sheet.get("grid_properties") or {}
            row_count = int(grid.get("row_count") or 50)
            col_count = int(grid.get("column_count") or 20)
            rows_to_read = min(row_count, int(selection.get("max_rows") or 200))
            range_name = selection.get("range") or f"{sheet_id}!A1:{self._column_name(min(col_count, 26))}{rows_to_read}"
            values_body = await self._request_json(
                "GET",
                f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{range_name}",
                access_token=access_token,
            )
            values = (((values_body.get("data") or {}).get("valueRange") or {}).get("values") or [])
            raw["sheets"].append({"sheet": sheet, "range": range_name, "values": values})
            lines.append(f"## Sheet: {title}\n\n{_table_markdown(values)}")
        return "\n\n".join(lines), raw, self._stable_revision(raw)

    async def _fetch_bitable(
        self,
        *,
        access_token: str,
        app_token: str,
        selection: dict[str, Any],
    ) -> tuple[str, dict[str, Any], str | None]:
        tables_body = await self._request_json("GET", f"/open-apis/bitable/v1/apps/{app_token}/tables", access_token=access_token)
        tables = ((tables_body.get("data") or {}).get("items") or [])
        selected = {item.get("table_id") for item in (selection.get("subresources") or []) if item.get("table_id")}
        raw = {"tables": []}
        lines = [f"# Bitable {app_token}"]
        for table in tables:
            table_id = table.get("table_id")
            if selected and table_id not in selected:
                continue
            fields_body = await self._request_json("GET", f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields", access_token=access_token)
            records_body = await self._request_json(
                "GET",
                f"/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records",
                access_token=access_token,
                params={"page_size": min(int(selection.get("max_records") or 200), 500)},
            )
            raw["tables"].append({"table": table, "fields": fields_body.get("data"), "records": records_body.get("data")})
            field_names = [field.get("field_name") for field in ((fields_body.get("data") or {}).get("items") or [])]
            rows = []
            for record in ((records_body.get("data") or {}).get("items") or []):
                fields = record.get("fields") or {}
                rows.append([fields.get(name, "") for name in field_names])
            lines.append(f"## {table.get('name') or table_id}\n\n{_table_markdown([field_names, *rows])}")
        return "\n\n".join(lines), raw, self._stable_revision(raw)

    def _extract_feishu_text(self, block: dict[str, Any]) -> str:
        values: list[str] = []
        def walk(value: Any) -> None:
            if isinstance(value, dict):
                if isinstance(value.get("text_run"), dict) and value["text_run"].get("content"):
                    values.append(str(value["text_run"]["content"]))
                elif isinstance(value.get("text"), str):
                    values.append(value["text"])
                for child in value.values():
                    walk(child)
            elif isinstance(value, list):
                for child in value:
                    walk(child)
        walk(block)
        return "".join(values).strip()

    def _stable_revision(self, value: Any) -> str:
        return "sha256:" + hashlib.sha256(
            json.dumps(value, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()

    def _column_name(self, index: int) -> str:
        index = max(1, index)
        result = ""
        while index:
            index, rem = divmod(index - 1, 26)
            result = chr(65 + rem) + result
        return result


class TosConnectorAdapter:
    provider = "volcengine_tos"
    max_inline_bytes = 20 * 1024 * 1024

    async def test_connection(self, credentials: dict[str, Any]) -> dict[str, Any]:
        client = self._client(credentials)
        try:
            output = client.list_buckets()
        except Exception as exc:
            raise ConnectorError(f"TOS connection failed: {exc}", code="connection_failed") from exc
        buckets = getattr(output, "buckets", None) or []
        return {
            "bucket_count": len(buckets),
            "default_bucket": credentials.get("default_bucket"),
            "region": credentials.get("region"),
            "endpoint": credentials.get("endpoint"),
        }

    async def list_resources(
        self,
        *,
        session: AsyncSession,
        input: ResourceListInput,
    ) -> ResourceListResult:
        credentials = await CryptoService.decrypt_config(input.connection.encrypted_credentials, session)
        client = self._client(credentials)
        default_bucket = credentials.get("default_bucket")
        if input.scope == "bucket" or not (input.parent_token or default_bucket):
            output = client.list_buckets()
            buckets = getattr(output, "buckets", None) or []
            items = [
                ResourcePickerItem(
                    external_id=getattr(bucket, "name", ""),
                    resource_type="tos_bucket",
                    name=getattr(bucket, "name", ""),
                    has_children=True,
                    is_folder=True,
                    already_added=getattr(bucket, "name", "") in input.already_added_external_ids,
                    metadata={"creation_date": str(getattr(bucket, "creation_date", ""))},
                )
                for bucket in buckets
            ]
            return ResourceListResult(items=items, next_page_token=None)

        bucket, prefix = self._split_external_id(input.parent_token or default_bucket)
        prefix = input.query or prefix
        output = client.list_objects_type2(
            bucket=bucket,
            prefix=prefix or credentials.get("default_prefix") or "",
            delimiter="/",
            continuation_token=input.page_token,
            max_keys=input.page_size,
        )
        folders = getattr(output, "common_prefixes", None) or []
        objects = getattr(output, "contents", None) or []
        items: list[ResourcePickerItem] = []
        for folder in folders:
            key = getattr(folder, "prefix", None) or (folder.get("prefix") if isinstance(folder, dict) else "")
            items.append(
                ResourcePickerItem(
                    external_id=f"{bucket}/{key}",
                    resource_type="tos_prefix",
                    name=key.rstrip("/").rsplit("/", 1)[-1] or key,
                    parent_external_id=bucket,
                    has_children=True,
                    is_folder=True,
                    already_added=f"{bucket}/{key}" in input.already_added_external_ids,
                    metadata={"bucket": bucket, "prefix": key},
                )
            )
        for obj in objects:
            key = getattr(obj, "key", "")
            if key.endswith("/"):
                continue
            etag = getattr(obj, "etag", None)
            external_id = f"{bucket}/{key}"
            items.append(
                ResourcePickerItem(
                    external_id=external_id,
                    resource_type="tos_object",
                    name=key.rsplit("/", 1)[-1],
                    parent_external_id=bucket,
                    has_children=False,
                    is_folder=False,
                    already_added=external_id in input.already_added_external_ids,
                    metadata={
                        "bucket": bucket,
                        "key": key,
                        "size": getattr(obj, "size", None),
                        "last_modified": str(getattr(obj, "last_modified", "")),
                        "etag": etag.strip('"') if isinstance(etag, str) else etag,
                        "storage_class": getattr(obj, "storage_class", None),
                    },
                )
            )
        return ResourceListResult(items=items, next_page_token=getattr(output, "next_continuation_token", None))

    async def sync_resource(
        self,
        *,
        session: AsyncSession,
        connection: SourceConnection,
        resource: SourceResource,
    ) -> CapturedSnapshot:
        credentials = await CryptoService.decrypt_config(connection.encrypted_credentials, session)
        if resource.resource_type == "tos_prefix":
            return await self._sync_prefix(credentials=credentials, resource=resource)
        if resource.resource_type == "tos_bucket":
            return await self._sync_prefix(credentials=credentials, resource=resource, bucket_only=True)
        return await self._sync_object(credentials=credentials, resource=resource)

    async def _sync_object(self, *, credentials: dict[str, Any], resource: SourceResource) -> CapturedSnapshot:
        client = self._client(credentials)
        bucket, key = self._split_external_id(resource.external_id or "")
        try:
            head = client.head_object(bucket=bucket, key=key)
        except Exception as exc:
            raise self._classify_tos_exception(exc) from exc
        size = int(getattr(head, "content_length", 0) or 0)
        if size > self.max_inline_bytes and not (resource.selection_config_json or {}).get("allow_large_file"):
            raise ConnectorError("TOS object is too large; confirmation required", code="large_file_confirmation_required", permanent=True)
        try:
            output = client.get_object(bucket=bucket, key=key)
            raw_bytes = output.read()
        except Exception as exc:
            raise self._classify_tos_exception(exc) from exc
        etag = getattr(head, "etag", None) or getattr(output, "etag", None)
        last_modified = getattr(head, "last_modified", None)
        version_id = getattr(head, "version_id", None) or getattr(output, "version_id", None)
        text, parser_version, fragment_hint = parse_object_bytes(key=key, raw_bytes=raw_bytes)
        metadata = {
            "provider": self.provider,
            "bucket": bucket,
            "key": key,
            "region": credentials.get("region"),
            "endpoint": credentials.get("endpoint"),
            "etag": etag.strip('"') if isinstance(etag, str) else etag,
            "last_modified": str(last_modified) if last_modified else None,
            "version_id": str(version_id) if version_id else None,
            "size": size,
            "fragment_hint": fragment_hint,
        }
        return CapturedSnapshot(
            raw_bytes=raw_bytes,
            content_text=text,
            external_revision=metadata["etag"] or metadata["last_modified"],
            metadata=metadata,
            provider="byaan-native",
            parser_version=parser_version,
            raw_storage_uri=f"tos://{bucket}/{key}",
        )

    async def _sync_prefix(
        self,
        *,
        credentials: dict[str, Any],
        resource: SourceResource,
        bucket_only: bool = False,
    ) -> CapturedSnapshot:
        client = self._client(credentials)
        bucket, prefix = self._split_external_id(resource.external_id or "")
        if bucket_only:
            prefix = ""
        output = client.list_objects_type2(bucket=bucket, prefix=prefix, max_keys=1000)
        contents = getattr(output, "contents", None) or []
        rows = []
        for obj in contents:
            key = getattr(obj, "key", "")
            if key.endswith("/"):
                continue
            rows.append(
                {
                    "key": key,
                    "size": getattr(obj, "size", None),
                    "etag": getattr(obj, "etag", None),
                    "last_modified": str(getattr(obj, "last_modified", "")),
                }
            )
        text = "\n".join(json.dumps(row, ensure_ascii=False, sort_keys=True) for row in rows)
        raw_bytes = json.dumps(rows, ensure_ascii=False, sort_keys=True).encode("utf-8")
        collection_revision = "collection:sha256:" + hashlib.sha256(raw_bytes).hexdigest()
        return CapturedSnapshot(
            raw_bytes=raw_bytes,
            content_text=text,
            external_revision=collection_revision,
            metadata={
                "provider": self.provider,
                "bucket": bucket,
                "prefix": prefix,
                "object_count": len(rows),
                "region": credentials.get("region"),
            },
            provider="byaan-native",
            parser_version="tos-prefix-listing-v1",
            raw_storage_uri=f"tos://{bucket}/{prefix}",
        )

    def _client(self, credentials: dict[str, Any]):
        try:
            import tos
        except ImportError as exc:
            raise ConnectorError("Volcengine TOS SDK is not installed", code="sdk_missing", permanent=True) from exc
        return tos.TosClientV2(
            ak=credentials.get("access_key_id") or credentials.get("ak") or "",
            sk=credentials.get("secret_access_key") or credentials.get("sk") or "",
            endpoint=credentials.get("endpoint") or "",
            region=credentials.get("region") or "",
            security_token=credentials.get("session_token"),
            enable_verify_ssl=credentials.get("verify_ssl", True),
        )

    def _split_external_id(self, external_id: str) -> tuple[str, str]:
        value = external_id.removeprefix("tos://")
        if "/" not in value:
            return value, ""
        bucket, key = value.split("/", 1)
        return bucket, key

    def _classify_tos_exception(self, exc: Exception) -> ConnectorError:
        text = str(exc).lower()
        if any(term in text for term in ("accessdenied", "access denied", "permission", "forbidden", "unauthorized")):
            return ConnectorError(f"TOS permission denied: {exc}", code="permission_lost", permanent=True)
        if any(term in text for term in ("invalidaccesskey", "signaturedoesnotmatch", "securitytoken", "auth", "credential")):
            return ConnectorError(f"TOS authorization failed: {exc}", code="authorization_required", permanent=True)
        if any(term in text for term in ("nosuchkey", "nosuchbucket", "not found", "not exist", "404")):
            return ConnectorError(f"TOS resource unavailable: {exc}", code="source_unavailable", permanent=True)
        if any(term in text for term in ("timeout", "timed out", "read timed")):
            return ConnectorError(f"TOS request timed out: {exc}", code="source_timeout")
        if any(term in text for term in ("rate", "too many", "throttl", "slow down", "qps")):
            return ConnectorError(f"TOS rate limited: {exc}", code="rate_limited")
        return ConnectorError(f"TOS request failed: {exc}", code="tos_request_failed")


def parse_object_bytes(*, key: str, raw_bytes: bytes) -> tuple[str, str, str]:
    suffix = key.rsplit(".", 1)[-1].lower() if "." in key else ""
    if suffix in {"txt", "md", "markdown", "log"}:
        return raw_bytes.decode("utf-8", errors="replace"), f"tos-{suffix}-parser-v1", "raw_text"
    if suffix == "csv":
        text = raw_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        return _table_markdown(rows[:500]), "tos-csv-parser-v1", "csv_rows"
    if suffix in {"json", "jsonl"}:
        text = raw_bytes.decode("utf-8", errors="replace")
        try:
            if suffix == "json":
                value = json.loads(text)
                return json.dumps(value, ensure_ascii=False, indent=2)[:100000], "tos-json-parser-v1", "json_records"
            records = [json.loads(line) for line in text.splitlines() if line.strip()]
            return (
                "\n".join(json.dumps(record, ensure_ascii=False, sort_keys=True) for record in records[:1000]),
                "tos-jsonl-parser-v1",
                "json_records",
            )
        except json.JSONDecodeError as exc:
            raise ConnectorError(f"Invalid JSON object content: {exc}", code="parser_invalid_json", permanent=True) from exc
    if suffix in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ConnectorError("openpyxl is required to parse Excel objects", code="parser_missing", permanent=True) from exc
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=False, read_only=True)
        parts = []
        for sheet in workbook.worksheets[:20]:
            rows = []
            for row in sheet.iter_rows(max_row=500, values_only=True):
                rows.append(["" if value is None else str(value) for value in row])
            parts.append(f"## Sheet: {sheet.title}\n\n{_table_markdown(rows)}")
        return "\n\n".join(parts), "tos-excel-parser-v1", "excel_range"
    if suffix == "parquet":
        try:
            import duckdb
        except ImportError as exc:
            raise ConnectorError("duckdb is required to parse Parquet objects", code="parser_missing", permanent=True) from exc
        with tempfile.NamedTemporaryFile(suffix=".parquet") as parquet_file:
            parquet_file.write(raw_bytes)
            parquet_file.flush()
            relation = duckdb.from_parquet(parquet_file.name)
            rows = relation.limit(500).fetchall()
            columns = [column[0] for column in relation.description]
        return _table_markdown([columns, *rows]), "tos-parquet-parser-v1", "parquet_rows"
    if suffix in {"html", "htm"}:
        text = raw_bytes.decode("utf-8", errors="replace")
        text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", text)
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        return html.unescape(re.sub(r"\s+", " ", text)).strip(), "tos-html-parser-v1", "html_section"
    if suffix == "docx":
        return _parse_docx_bytes(raw_bytes), "tos-docx-parser-v1", "docx_paragraph"
    if suffix == "pdf":
        return _parse_pdf_text_fallback(raw_bytes), "tos-pdf-basic-parser-v1", "page"
    raise ConnectorError(f"Unsupported TOS object format: .{suffix or 'unknown'}", code="unsupported_format", permanent=True)


def _parse_docx_bytes(raw_bytes: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
    texts = re.findall(r"<w:t[^>]*>(.*?)</w:t>", document_xml)
    return "\n".join(html.unescape(re.sub(r"<[^>]+>", "", text)) for text in texts if text)


def _parse_pdf_text_fallback(raw_bytes: bytes) -> str:
    text = raw_bytes.decode("latin-1", errors="ignore")
    chunks = re.findall(r"\(([^()]{2,})\)\s*Tj", text)
    if not chunks:
        chunks = re.findall(r"\(([^()]{2,})\)", text)
    extracted = "\n".join(html.unescape(chunk) for chunk in chunks[:500]).strip()
    if not extracted:
        raise ConnectorError("PDF text extraction produced no text; configure a PDF parser worker", code="parser_no_text")
    return extracted


def _table_markdown(rows: list[Any]) -> str:
    if not rows:
        return ""
    normalized = [[str(cell) if cell is not None else "" for cell in row] for row in rows]
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    header = padded[0]
    body = padded[1:]
    lines = [
        "| " + " | ".join(_escape_table_cell(cell) for cell in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body[:500]:
        lines.append("| " + " | ".join(_escape_table_cell(cell) for cell in row) + " |")
    return "\n".join(lines)


def _escape_table_cell(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


def get_connector_adapter(provider: str) -> SourceConnectorAdapter:
    if provider == "feishu":
        return FeishuConnectorAdapter()
    if provider == "volcengine_tos":
        return TosConnectorAdapter()
    raise ConnectorError(f"Unsupported connector provider: {provider}", code="unsupported_provider", permanent=True)
